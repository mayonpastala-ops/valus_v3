#!/usr/bin/env python3
"""
Valus DCF — Institutional-quality discounted cash flow model.
Usage: python3 dcf.py AAPL
"""

import sys
import warnings
from datetime import date

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL STYLES
# ═══════════════════════════════════════════════════════════════════════════════

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BOLD = Font(bold=True, size=10, name="Calibri")
NORM = Font(size=10, name="Calibri")
TITLE = Font(bold=True, size=13, color="1F3864", name="Calibri")
SUB = Font(bold=True, size=11, color="1F3864", name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREEN_FNT = Font(bold=True, size=11, color="006100", name="Calibri")
YELLOW_FNT = Font(bold=True, size=11, color="9C6500", name="Calibri")
RED_FNT = Font(bold=True, size=11, color="9C0006", name="Calibri")
THIN = Border(bottom=Side("thin", "C0C0C0"))
THICK_BOTTOM = Border(bottom=Side("medium", "1F3864"))
DOLLAR = '#,##0'
DOLLAR2 = '$#,##0.00'
DOLLAR0 = '$#,##0'
PCT = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0'
MULT = '0.0x'

# ── Currency globals (set per-run by _init_currency) ──────────────────────────
_CUR_SYMBOL = '$'
_CUR_CODE = 'USD'
_CUR_LABEL = '$'  # for sheet titles

_CURRENCY_SYMBOLS = {
    'USD': '$', 'INR': '\u20B9', 'GBP': '\u00A3', 'EUR': '\u20AC',
    'HKD': 'HK$', 'AUD': 'A$', 'CAD': 'C$', 'JPY': '\u00A5',
    'CNY': '\u00A5', 'KRW': '\u20A9', 'CHF': 'CHF ', 'SEK': 'kr',
}


def _init_currency(code):
    """Set module-level currency formats based on stock's reporting currency."""
    global _CUR_SYMBOL, _CUR_CODE, _CUR_LABEL, DOLLAR2, DOLLAR0
    _CUR_CODE = code or 'USD'
    _CUR_SYMBOL = _CURRENCY_SYMBOLS.get(_CUR_CODE, _CUR_CODE + ' ')
    _CUR_LABEL = _CUR_SYMBOL.strip()
    DOLLAR2 = f'{_CUR_SYMBOL}#,##0.00'
    DOLLAR0 = f'{_CUR_SYMBOL}#,##0'


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _v(series, labels, default=0.0):
    """Extract value from a pandas Series, trying multiple row labels."""
    if series is None:
        return default
    for lab in (labels if isinstance(labels, list) else [labels]):
        if lab in series.index:
            val = series[lab]
            if pd.notna(val):
                return float(val)
    return default


def _pct(num, denom):
    """Safe percentage."""
    if denom and denom != 0:
        return num / denom
    return 0.0


def _avg(lst):
    """Average of non-zero values."""
    vals = [v for v in lst if v and v != 0]
    return sum(vals) / len(vals) if vals else 0.0


def _cagr(first, last, years):
    """Compound annual growth rate."""
    if first and first > 0 and last and last > 0 and years > 0:
        return (last / first) ** (1 / years) - 1
    return 0.0


def _hdr(ws, row, values):
    """Write a styled header row."""
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")


def _alt_rows(ws, r_start, r_end, max_col):
    """Apply alternating row shading."""
    for r in range(r_start, r_end + 1):
        fill = ALT_FILL if (r - r_start) % 2 == 0 else WHITE_FILL
        for c in range(1, max_col + 1):
            ws.cell(r, c).fill = fill
            ws.cell(r, c).border = THIN


def _autowidth(ws, mn=13, mx=24):
    """Auto-fit column widths."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[letter].width = max(min(best + 3, mx), mn)


def _lv(ws, row, label, value, fmt=None):
    """Write label-value pair in columns A-B."""
    ws.cell(row, 1, label).font = BOLD
    c = ws.cell(row, 2, value)
    c.font = NORM
    if fmt:
        c.number_format = fmt
    return row + 1


# ═══════════════════════════════════════════════════════════════════════════════
# DATA FETCHING
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_all(ticker):
    """Fetch all historical data for a ticker. Returns (hist_dict, stock, warns)."""
    stock = yf.Ticker(ticker)
    inc = stock.financials
    bs = stock.balance_sheet
    cf = stock.cashflow
    warns = []

    # Allow partial data — only need at least one non-empty statement
    has_inc = not inc.empty if inc is not None else False
    has_bs = not bs.empty if bs is not None else False
    has_cf = not cf.empty if cf is not None else False

    if not has_inc and not has_bs and not has_cf:
        raise ValueError(f"No financial data available for {ticker}")

    # Use empty DataFrames as fallback for missing statements
    if not has_inc:
        warns.append("Income statement unavailable — using zeros")
        inc = pd.DataFrame()
    if not has_bs:
        warns.append("Balance sheet unavailable — using zeros")
        bs = pd.DataFrame()
    if not has_cf:
        warns.append("Cash flow statement unavailable — using zeros")
        cf = pd.DataFrame()

    # Determine number of years from whichever statement has data
    col_counts = []
    if has_inc:
        col_counts.append(len(inc.columns))
    if has_bs:
        col_counts.append(len(bs.columns))
    if has_cf:
        col_counts.append(len(cf.columns))
    n = min(min(col_counts), 4) if col_counts else 0
    if n < 2:
        raise ValueError(f"Need >= 2 years of data, got {n}")

    h = {
        "years": [], "revenue": [], "cogs": [], "gross_profit": [],
        "sga": [], "da": [], "ebit": [], "interest": [], "ebt": [],
        "tax": [], "net_income": [], "ebitda": [],
        "cash": [], "receivables": [], "inventory": [], "prepaid": [],
        "total_ca": [], "ppe_gross": [], "accum_depr": [], "ppe_net": [],
        "goodwill": [], "intangibles": [], "total_assets": [],
        "payables": [], "accrued": [], "current_debt": [],
        "total_cl": [], "lt_debt": [], "lease_liab": [],
        "total_liab": [], "common_equity": [], "retained_earnings": [],
        "total_equity": [],
        "cf_ni": [], "cf_da": [], "sbc": [], "wc_change": [], "cfo": [],
        "capex": [], "cfi": [], "debt_issued": [], "debt_repaid": [],
        "equity_issued": [], "equity_repurchased": [], "cff": [],
        "net_cash_change": [], "end_cash": [],
    }

    # Track which fields came back all-zero so we can warn
    _zero_fields = set()

    for i in range(n - 1, -1, -1):  # oldest first
        # Get column for each statement (may be different lengths)
        ci = inc.columns[i] if has_inc and i < len(inc.columns) else None
        cb = bs.columns[i] if has_bs and i < len(bs.columns) else None
        cc = cf.columns[i] if has_cf and i < len(cf.columns) else None

        yr_source = ci or cb or cc
        yr = yr_source.year if hasattr(yr_source, "year") else int(str(yr_source)[:4])
        h["years"].append(yr)

        si = inc[ci] if ci is not None else None
        sb = bs[cb] if cb is not None else None
        sc = cf[cc] if cc is not None else None

        # ── Income statement (expanded name variants) ─────────────────
        h["revenue"].append(_v(si, ["Total Revenue", "Operating Revenue", "Revenue"]))
        h["cogs"].append(_v(si, ["Cost Of Revenue", "Reconciled Cost Of Revenue",
                                  "Cost Of Goods Sold"]))
        h["gross_profit"].append(_v(si, ["Gross Profit"]))
        h["sga"].append(_v(si, ["Selling General And Administration",
                                  "General And Administrative Expense",
                                  "Operating Expense"]))
        h["da"].append(_v(sc, ["Depreciation And Amortization",
                                "Depreciation & Amortization",
                                "Depreciation"]) or
                       _v(si, ["Reconciled Depreciation"]))
        h["ebit"].append(_v(si, ["EBIT", "Operating Income"]))
        h["interest"].append(_v(si, ["Interest Expense",
                                      "Interest Expense Non Operating",
                                      "Net Interest Income"]))
        h["ebt"].append(_v(si, ["Pretax Income", "Income Before Tax"]))
        h["tax"].append(_v(si, ["Tax Provision", "Income Tax Expense"]))
        h["net_income"].append(_v(si, ["Net Income",
                                        "Net Income Common Stockholders",
                                        "Net Income Including Noncontrolling Interests"]))
        h["ebitda"].append(_v(si, ["EBITDA", "Normalized EBITDA"]))

        # ── Balance sheet (expanded name variants) ────────────────────
        h["cash"].append(_v(sb, ["Cash And Cash Equivalents",
                                  "Cash Cash Equivalents And Short Term Investments",
                                  "Cash Financial", "Cash Equivalents"]))
        h["receivables"].append(_v(sb, ["Net Receivable", "Accounts Receivable",
                                         "Other Receivables"]))
        h["inventory"].append(_v(sb, ["Inventory"]))
        h["prepaid"].append(_v(sb, ["Other Current Assets", "Prepaid Assets"]))
        h["total_ca"].append(_v(sb, ["Current Assets"]))
        h["ppe_gross"].append(_v(sb, ["Gross Property Plant And Equipment",
                                       "Gross PPE"]))
        h["accum_depr"].append(_v(sb, ["Accumulated Depreciation"]))
        h["ppe_net"].append(_v(sb, ["Net Property Plant And Equipment",
                                     "Property Plant Equipment",
                                     "Net PPE"]))
        h["goodwill"].append(_v(sb, ["Goodwill"]))
        h["intangibles"].append(_v(sb, ["Intangible Assets",
                                         "Other Intangible Assets",
                                         "Goodwill And Other Intangible Assets"]))
        h["total_assets"].append(_v(sb, ["Total Assets"]))
        h["payables"].append(_v(sb, ["Accounts Payable", "Payables"]))
        h["accrued"].append(_v(sb, ["Other Current Liabilities",
                                     "Accrued Expense", "Current Provisions"]))
        h["current_debt"].append(_v(sb, ["Current Debt",
                                          "Current Portion Of Long Term Debt",
                                          "Current Debt And Capital Lease Obligation"]))
        h["total_cl"].append(_v(sb, ["Current Liabilities"]))
        h["lt_debt"].append(_v(sb, ["Long Term Debt",
                                     "Long Term Debt And Capital Lease Obligation"]))
        h["lease_liab"].append(_v(sb, ["Operating Lease Liability",
                                        "Long Term Operating Lease",
                                        "Long Term Capital Lease Obligation",
                                        "Capital Lease Obligations"]))
        h["total_liab"].append(_v(sb, ["Total Liabilities Net Minority Interest",
                                        "Total Non Current Liabilities Net Minority Interest"]))
        h["common_equity"].append(_v(sb, ["Common Stock Equity",
                                           "Stockholders Equity"]))
        h["retained_earnings"].append(_v(sb, ["Retained Earnings"]))
        h["total_equity"].append(_v(sb, ["Total Equity Gross Minority Interest",
                                          "Stockholders Equity"]))

        # ── Cash flow (expanded name variants) ────────────────────────
        h["cf_ni"].append(_v(sc, ["Net Income",
                                   "Net Income From Continuing Operations",
                                   "Net Income Continuous Operations"]))
        h["cf_da"].append(_v(sc, ["Depreciation And Amortization",
                                   "Depreciation & Amortization",
                                   "Depreciation"]))
        h["sbc"].append(_v(sc, ["Stock Based Compensation"]))
        h["wc_change"].append(_v(sc, ["Change In Working Capital"]))
        h["cfo"].append(_v(sc, ["Operating Cash Flow"]))
        h["capex"].append(_v(sc, ["Capital Expenditure"]))
        h["cfi"].append(_v(sc, ["Investing Cash Flow", "Investing Activities"]))
        h["debt_issued"].append(_v(sc, ["Issuance Of Debt",
                                         "Long Term Debt Issuance"]))
        h["debt_repaid"].append(_v(sc, ["Repayment Of Debt",
                                         "Long Term Debt Payments"]))
        h["equity_issued"].append(_v(sc, ["Common Stock Issuance",
                                           "Issuance Of Capital Stock"]))
        h["equity_repurchased"].append(_v(sc, ["Repurchase Of Capital Stock",
                                                "Common Stock Payments"]))
        h["cff"].append(_v(sc, ["Financing Cash Flow", "Financing Activities"]))
        h["net_cash_change"].append(_v(sc, ["Changes In Cash",
                                             "Change In Cash And Cash Equivalents"]))
        h["end_cash"].append(_v(sc, ["End Cash Position"]))

    # Check for all-zero fields and log warnings
    _skip = {"years"}
    for key, vals in h.items():
        if key in _skip or not isinstance(vals, list):
            continue
        if all(v == 0 or v == 0.0 for v in vals):
            _zero_fields.add(key)

    _field_labels = {
        "ppe_net": "PP&E (Net)", "ppe_gross": "PP&E (Gross)",
        "intangibles": "Intangible Assets", "lease_liab": "Lease Liabilities",
        "cf_ni": "CF Net Income", "sbc": "Stock-Based Compensation",
        "equity_repurchased": "Share Buybacks", "accum_depr": "Accumulated Depreciation",
        "goodwill": "Goodwill", "prepaid": "Prepaid Assets",
    }
    for f in sorted(_zero_fields):
        label = _field_labels.get(f, f)
        warns.append(f"{label} unavailable — using zero")

    return h, stock, warns


def fetch_risk_free():
    """10-year US Treasury yield from ^TNX."""
    try:
        tnx = yf.Ticker("^TNX")
        hist = tnx.history(period="5d")
        if not hist.empty:
            rate = hist["Close"].iloc[-1] / 100
            if 0.005 < rate < 0.20:  # sanity check
                return rate
    except Exception:
        pass
    return 0.045  # fallback


def _is_indian(ticker):
    return ticker.upper().endswith((".NS", ".BO"))


# ═══════════════════════════════════════════════════════════════════════════════
# WACC
# ═══════════════════════════════════════════════════════════════════════════════

def calc_wacc(h, stock):
    """Calculate WACC from scratch. Returns (wacc_float, components_dict)."""
    info = stock.info
    ticker = info.get("symbol", "")
    rf = fetch_risk_free()

    # Beta — fall back to 1.0 if unavailable
    beta = info.get("beta")
    if not beta or beta <= 0:
        beta = 1.0
    erp = 0.055
    ke = rf + beta * erp  # CAPM

    # Cost of debt — with regional fallback
    interest = abs(h["interest"][-1]) if h["interest"][-1] else 0
    total_debt = (h["lt_debt"][-1] or 0) + (h["current_debt"][-1] or 0)

    # Use Total Debt from balance sheet if lt_debt + current_debt is zero
    if total_debt <= 0:
        # Try yfinance info
        total_debt = info.get("totalDebt", 0) or 0
    if total_debt <= 0:
        total_debt = 1  # avoid division by zero

    if interest > 0 and total_debt > 1:
        kd = interest / total_debt
    else:
        # Regional fallback
        kd = 0.07 if _is_indian(ticker) else 0.05

    # Effective tax rate
    tax_prov = h["tax"][-1]
    ebt = h["ebt"][-1]
    if ebt and ebt > 0 and tax_prov:
        tax_rate = min(max(abs(tax_prov) / ebt, 0), 0.40)
    else:
        # Regional defaults
        tax_rate = 0.25 if _is_indian(ticker) else 0.21
    kd_at = kd * (1 - tax_rate)

    mktcap = info.get("marketCap") or 0
    if mktcap <= 0:
        price = info.get("currentPrice") or info.get("regularMarketPrice") or 1
        shares = info.get("sharesOutstanding") or 1
        mktcap = price * shares

    total_cap = mktcap + total_debt
    we = mktcap / total_cap
    wd = total_debt / total_cap
    wacc = we * ke + wd * kd_at

    comp = dict(rf=rf, beta=beta, erp=erp, ke=ke, interest=interest,
                total_debt=total_debt, kd=kd, tax_rate=tax_rate, kd_at=kd_at,
                mktcap=mktcap, total_cap=total_cap, we=we, wd=wd, wacc=wacc)
    return wacc, comp


# ═══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE SCENARIO PROMPTS
# ═══════════════════════════════════════════════════════════════════════════════

def _ask(prompt, default):
    """Prompt user for a float value with a default."""
    raw = input(f"  {prompt} [{default}]: ").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        print(f"    Invalid input, using {default}")
        return default


def get_scenarios(h, wacc):
    """Interactively gather Base/Bull/Bear assumptions."""
    # Calculate historical revenue CAGR as a reference
    revs = [r for r in h["revenue"] if r and r > 0]
    hist_cagr = _cagr(revs[0], revs[-1], len(revs) - 1) if len(revs) >= 2 else 0.05

    print(f"\n  Historical revenue CAGR: {hist_cagr:.1%}")
    print(f"  Calculated WACC: {wacc:.2%}")
    print()
    print("  Enter assumptions (press Enter for defaults):\n")

    scenarios = {}
    defaults = {
        "Base": (round(hist_cagr * 100, 1), round(wacc * 100, 1), 2.5),
        "Bull": (round(hist_cagr * 100 * 1.5, 1), round((wacc - 0.01) * 100, 1), 3.0),
        "Bear": (round(max(hist_cagr * 100 * 0.5, 1), 1), round((wacc + 0.01) * 100, 1), 2.0),
    }

    for name in ["Base", "Bull", "Bear"]:
        dg, dw, dt = defaults[name]
        print(f"  ── {name} Case ──")
        g = _ask(f"Revenue growth % (annual)", dg)
        w = _ask(f"WACC %", dw)
        t = _ask(f"Terminal growth %", dt)
        scenarios[name] = dict(rev_growth=g / 100, wacc=w / 100, tgr=t / 100)
        print()

    return scenarios


# ═══════════════════════════════════════════════════════════════════════════════
# PROJECTION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def build_projections(h, scenario, n_years=5):
    """Build full 5-year projections. Returns dict mirroring hist structure."""
    g = scenario["rev_growth"]
    g = max(-0.05, min(0.40, g))

    # Historical driver ratios (averages)
    revs = h["revenue"]
    cogs_pct = _avg([_pct(h["cogs"][i], revs[i]) for i in range(len(revs))])
    sga_pct = _avg([_pct(h["sga"][i], revs[i]) for i in range(len(revs))])
    da_pct_ppe = _avg([_pct(h["da"][i], h["ppe_net"][i - 1] if i > 0 else h["ppe_net"][i])
                        for i in range(len(revs)) if h["ppe_net"][i]])
    if da_pct_ppe == 0:
        da_pct_ppe = _avg([_pct(h["da"][i], revs[i]) for i in range(len(revs))])
    capex_pct = _avg([_pct(abs(h["capex"][i]), revs[i]) for i in range(len(revs))])
    int_pct = _avg([_pct(abs(h["interest"][i]), revs[i]) for i in range(len(revs))])

    # NWC driver ratios
    ar_pct = _avg([_pct(h["receivables"][i], revs[i]) for i in range(len(revs))])
    inv_pct = _avg([_pct(h["inventory"][i], revs[i]) for i in range(len(revs))])
    pre_pct = _avg([_pct(h["prepaid"][i], revs[i]) for i in range(len(revs))])
    ap_pct = _avg([_pct(h["payables"][i], revs[i]) for i in range(len(revs))])
    acc_pct = _avg([_pct(h["accrued"][i], revs[i]) for i in range(len(revs))])

    tax_rate = scenario.get("tax_rate", 0.21)
    # Try to get effective tax rate from historicals
    eff_taxes = [_pct(abs(h["tax"][i]), h["ebt"][i])
                 for i in range(len(revs)) if h["ebt"][i] and h["ebt"][i] > 0]
    if eff_taxes:
        tax_rate = min(_avg(eff_taxes), 0.35)

    last_yr = h["years"][-1]
    p = {
        "years": [last_yr + i + 1 for i in range(n_years)],
        "revenue": [], "cogs": [], "gross_profit": [], "sga": [],
        "da": [], "ebit": [], "interest": [], "ebt": [], "tax": [],
        "net_income": [],
        # PPE
        "ppe_begin": [], "capex": [], "ppe_end": [],
        # NWC
        "receivables": [], "inventory": [], "prepaid": [],
        "payables": [], "accrued": [], "nwc": [], "delta_nwc": [],
        # NWC drivers
        "ar_pct": ar_pct, "inv_pct": inv_pct, "pre_pct": pre_pct,
        "ap_pct": ap_pct, "acc_pct": acc_pct,
        # DCF
        "nopat": [], "ufcf": [],
        # Drivers stored for reference
        "rev_growth": g, "cogs_pct": cogs_pct, "sga_pct": sga_pct,
        "da_pct_ppe": da_pct_ppe, "capex_pct": capex_pct, "tax_rate": tax_rate,
    }

    prev_rev = revs[-1]
    prev_ppe = h["ppe_net"][-1] if h["ppe_net"][-1] else 0
    prev_nwc = ((h["receivables"][-1] or 0) + (h["inventory"][-1] or 0) +
                (h["prepaid"][-1] or 0) - (h["payables"][-1] or 0) -
                (h["accrued"][-1] or 0))

    for yr in range(n_years):
        rev = prev_rev * (1 + g)
        cogs = rev * cogs_pct
        gp = rev - cogs

        # PPE & D&A
        ppe_beg = prev_ppe
        capex = rev * capex_pct
        da = ppe_beg * da_pct_ppe if da_pct_ppe else rev * 0.03
        ppe_end = ppe_beg + capex - da

        sga = rev * sga_pct
        ebit = gp - sga - da
        interest = rev * int_pct
        ebt = ebit - interest
        tax = ebt * tax_rate if ebt > 0 else 0
        ni = ebt - tax

        # NWC
        ar = rev * ar_pct
        inv = rev * inv_pct
        pre = rev * pre_pct
        ap = rev * ap_pct
        acc = rev * acc_pct
        nwc = ar + inv + pre - ap - acc
        d_nwc = nwc - prev_nwc

        # Unlevered FCF = NOPAT + D&A - Capex - ΔWC
        nopat = ebit * (1 - tax_rate)
        ufcf = nopat + da - capex - d_nwc

        p["revenue"].append(rev)
        p["cogs"].append(cogs)
        p["gross_profit"].append(gp)
        p["sga"].append(sga)
        p["da"].append(da)
        p["ebit"].append(ebit)
        p["interest"].append(interest)
        p["ebt"].append(ebt)
        p["tax"].append(tax)
        p["net_income"].append(ni)
        p["ppe_begin"].append(ppe_beg)
        p["capex"].append(capex)
        p["ppe_end"].append(ppe_end)
        p["receivables"].append(ar)
        p["inventory"].append(inv)
        p["prepaid"].append(pre)
        p["payables"].append(ap)
        p["accrued"].append(acc)
        p["nwc"].append(nwc)
        p["delta_nwc"].append(d_nwc)
        p["nopat"].append(nopat)
        p["ufcf"].append(ufcf)

        prev_rev = rev
        prev_ppe = ppe_end
        prev_nwc = nwc

    return p


def run_valuation(proj, scenario, shares):
    """Run DCF valuation on projections. Returns dict of results."""
    wacc = scenario["wacc"]
    tgr = scenario["tgr"]

    pv_fcfs = []
    for i, fcf in enumerate(proj["ufcf"]):
        df = 1 / (1 + wacc) ** (i + 1)
        pv_fcfs.append({"year": i + 1, "fcf": fcf, "df": df, "pv": fcf * df})

    sum_pv = sum(p["pv"] for p in pv_fcfs)

    # Terminal value
    last_fcf = proj["ufcf"][-1]
    term_fcf = last_fcf * (1 + tgr)
    tv = term_fcf / (wacc - tgr) if wacc > tgr else term_fcf / 0.001
    pv_tv = tv / (1 + wacc) ** len(proj["ufcf"])

    ev = sum_pv + pv_tv
    net_debt = (proj.get("total_debt", 0) or 0) - (proj.get("cash", 0) or 0)
    equity = ev - net_debt
    ivps = equity / shares if shares > 0 else 0

    return dict(pv_fcfs=pv_fcfs, sum_pv=sum_pv, term_fcf=term_fcf,
                tv=tv, pv_tv=pv_tv, ev=ev, net_debt=net_debt,
                equity=equity, shares=shares, ivps=round(ivps, 2))


def sensitivity_table(base_val, proj, shares, base_wacc, n=5):
    """Build 5x5 sensitivity grid: WACC vs terminal growth."""
    wacc_steps = [base_wacc + (i - 2) * 0.005 for i in range(n)]
    tgr_steps = [0.015 + i * 0.005 for i in range(n)]  # 1.5% to 3.5%

    grid = []
    for w in wacc_steps:
        row = []
        for t in tgr_steps:
            # Recalculate terminal value with these assumptions
            pv_fcfs_sum = sum(
                fcf / (1 + w) ** (i + 1) for i, fcf in enumerate(proj["ufcf"])
            )
            last_fcf = proj["ufcf"][-1]
            tf = last_fcf * (1 + t)
            tv = tf / (w - t) if w > t else tf / 0.001
            pvtv = tv / (1 + w) ** len(proj["ufcf"])
            ev = pv_fcfs_sum + pvtv
            nd = base_val["net_debt"]
            eq = ev - nd
            ivps = eq / shares if shares > 0 else 0
            row.append(round(ivps, 2))
        grid.append(row)

    return wacc_steps, tgr_steps, grid


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def _col_labels(h, p):
    """Build column labels: '2021A', '2022A', '2025E', '2026E' etc."""
    hist_labels = [f"{y}A" for y in h["years"]]
    proj_labels = [f"{y}E" for y in p["years"]]
    return hist_labels, proj_labels


def _write_fs_row(ws, row, label, hist_vals, proj_vals, fmt, is_margin=False):
    """Write one financial statement row with hist + proj values."""
    ws.cell(row, 1, label).font = BOLD if not is_margin else NORM
    if is_margin:
        ws.cell(row, 1).font = Font(italic=True, size=9, color="666666", name="Calibri")
    for j, v in enumerate(hist_vals):
        c = ws.cell(row, j + 2, v)
        c.number_format = fmt
        c.font = NORM
    offset = len(hist_vals) + 2
    for j, v in enumerate(proj_vals):
        c = ws.cell(row, offset + j, v)
        c.number_format = fmt
        c.font = NORM
        c.fill = INPUT_FILL if not is_margin else WHITE_FILL


def sheet_income(wb, h, p):
    ws = wb.active
    ws.title = "Income Statement"
    hl, pl = _col_labels(h, p)
    ncols = 1 + len(hl) + len(pl)

    ws.cell(1, 1, f"Income Statement ({_CUR_LABEL} thousands)").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    _hdr(ws, 3, [""] + hl + pl)

    lines = [
        ("Revenue", "revenue", DOLLAR),
        ("  % Growth", None, PCT),
        ("Cost of Goods Sold", "cogs", DOLLAR),
        ("Gross Profit", "gross_profit", DOLLAR),
        ("  Gross Margin", None, PCT),
        ("SG&A", "sga", DOLLAR),
        ("Depreciation & Amortisation", "da", DOLLAR),
        ("EBIT", "ebit", DOLLAR),
        ("  EBIT Margin", None, PCT),
        ("Interest Expense", "interest", DOLLAR),
        ("EBT", "ebt", DOLLAR),
        ("Tax", "tax", DOLLAR),
        ("Net Income", "net_income", DOLLAR),
        ("  Net Margin", None, PCT),
    ]

    r = 4
    for label, key, fmt in lines:
        if key:
            hv = h[key]
            pv = p[key]
            _write_fs_row(ws, r, label, hv, pv, fmt)
        elif "Growth" in label:
            # Revenue growth
            hg = [0] + [_pct(h["revenue"][i] - h["revenue"][i - 1], h["revenue"][i - 1])
                        for i in range(1, len(h["revenue"]))]
            pg = [p["rev_growth"]] * len(p["revenue"])
            _write_fs_row(ws, r, label, hg, pg, PCT, is_margin=True)
        elif "Gross Margin" in label:
            hm = [_pct(h["gross_profit"][i], h["revenue"][i]) for i in range(len(h["revenue"]))]
            pm = [_pct(p["gross_profit"][i], p["revenue"][i]) for i in range(len(p["revenue"]))]
            _write_fs_row(ws, r, label, hm, pm, PCT, is_margin=True)
        elif "EBIT Margin" in label:
            hm = [_pct(h["ebit"][i], h["revenue"][i]) for i in range(len(h["revenue"]))]
            pm = [_pct(p["ebit"][i], p["revenue"][i]) for i in range(len(p["revenue"]))]
            _write_fs_row(ws, r, label, hm, pm, PCT, is_margin=True)
        elif "Net Margin" in label:
            hm = [_pct(h["net_income"][i], h["revenue"][i]) for i in range(len(h["revenue"]))]
            pm = [_pct(p["net_income"][i], p["revenue"][i]) for i in range(len(p["revenue"]))]
            _write_fs_row(ws, r, label, hm, pm, PCT, is_margin=True)
        r += 1

    # Thick line under headers for EBIT, Gross Profit, Net Income
    _alt_rows(ws, 4, r - 1, ncols)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_balance(wb, h):
    ws = wb.create_sheet("Balance Sheet")
    hl = [f"{y}A" for y in h["years"]]
    ncols = 1 + len(hl)

    ws.cell(1, 1, f"Balance Sheet ({_CUR_LABEL} thousands)").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _hdr(ws, 3, [""] + hl)

    lines = [
        ("ASSETS", None), ("", None),
        ("Cash & Equivalents", "cash"),
        ("Accounts Receivable", "receivables"),
        ("Inventory", "inventory"),
        ("Prepaid & Other Current", "prepaid"),
        ("Total Current Assets", "total_ca"),
        ("", None),
        ("PP&E (Gross)", "ppe_gross"),
        ("Accumulated Depreciation", "accum_depr"),
        ("PP&E (Net)", "ppe_net"),
        ("Goodwill", "goodwill"),
        ("Intangible Assets", "intangibles"),
        ("Total Assets", "total_assets"),
        ("", None),
        ("LIABILITIES", None), ("", None),
        ("Accounts Payable", "payables"),
        ("Accrued Expenses", "accrued"),
        ("Current Portion of Debt/Leases", "current_debt"),
        ("Total Current Liabilities", "total_cl"),
        ("", None),
        ("Long-Term Debt", "lt_debt"),
        ("Operating Lease Liabilities", "lease_liab"),
        ("Total Liabilities", "total_liab"),
        ("", None),
        ("EQUITY", None), ("", None),
        ("Common Equity", "common_equity"),
        ("Retained Earnings", "retained_earnings"),
        ("Total Equity", "total_equity"),
    ]

    r = 4
    for label, key in lines:
        if key is None:
            if label:
                ws.cell(r, 1, label).font = SUB
            r += 1
            continue
        ws.cell(r, 1, label).font = BOLD
        for j, v in enumerate(h[key]):
            ws.cell(r, j + 2, v).number_format = DOLLAR
        r += 1

    _alt_rows(ws, 4, r - 1, ncols)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_cashflow(wb, h):
    ws = wb.create_sheet("Cash Flow Statement")
    hl = [f"{y}A" for y in h["years"]]
    ncols = 1 + len(hl)

    ws.cell(1, 1, f"Cash Flow Statement ({_CUR_LABEL} thousands)").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _hdr(ws, 3, [""] + hl)

    lines = [
        ("OPERATING ACTIVITIES", None),
        ("Net Income", "cf_ni"),
        ("Depreciation & Amortisation", "cf_da"),
        ("Stock-Based Compensation", "sbc"),
        ("Change in Working Capital", "wc_change"),
        ("Cash from Operations (CFO)", "cfo"),
        ("", None),
        ("INVESTING ACTIVITIES", None),
        ("Capital Expenditure", "capex"),
        ("Cash from Investing (CFI)", "cfi"),
        ("", None),
        ("FINANCING ACTIVITIES", None),
        ("Debt Issuance", "debt_issued"),
        ("Debt Repayment", "debt_repaid"),
        ("Equity Issuance", "equity_issued"),
        ("Share Buybacks", "equity_repurchased"),
        ("Cash from Financing (CFF)", "cff"),
        ("", None),
        ("Net Change in Cash", "net_cash_change"),
        ("Ending Cash Balance", "end_cash"),
    ]

    r = 4
    for label, key in lines:
        if key is None:
            if label:
                ws.cell(r, 1, label).font = SUB
            r += 1
            continue
        ws.cell(r, 1, label).font = BOLD
        for j, v in enumerate(h[key]):
            ws.cell(r, j + 2, v).number_format = DOLLAR
        r += 1

    _alt_rows(ws, 4, r - 1, ncols)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_dcf(wb, h, p, val, wacc_comp, sens_data, scenario):
    ws = wb.create_sheet("DCF")
    ws.cell(1, 1, "Unlevered DCF Valuation").font = TITLE
    ws.merge_cells("A1:G1")

    # ── Projected FCF table ───────────────────────────────────────────────
    r = 3
    ws.cell(r, 1, "Projected Unlevered Free Cash Flow").font = SUB
    r += 1
    cols = [""] + [f"Year {v['year']}" for v in val["pv_fcfs"]]
    _hdr(ws, r, cols)

    items = [
        ("NOPAT", p["nopat"], DOLLAR),
        ("+ D&A", p["da"], DOLLAR),
        ("- Capex", p["capex"], DOLLAR),
        ("- Δ NWC", p["delta_nwc"], DOLLAR),
        ("Unlevered FCF", p["ufcf"], DOLLAR),
    ]
    for label, vals, fmt in items:
        r += 1
        ws.cell(r, 1, label).font = BOLD
        for j, v in enumerate(vals):
            ws.cell(r, j + 2, v).number_format = fmt

    _alt_rows(ws, 5, r, len(cols))

    # ── Discount table ────────────────────────────────────────────────────
    r += 2
    ws.cell(r, 1, "Discounted Cash Flows").font = SUB
    r += 1
    _hdr(ws, r, ["", "FCF", "Discount Factor", "PV of FCF"])
    for pf in val["pv_fcfs"]:
        r += 1
        ws.cell(r, 1, f"Year {pf['year']}").font = BOLD
        ws.cell(r, 2, pf["fcf"]).number_format = DOLLAR
        ws.cell(r, 3, pf["df"]).number_format = '0.0000'
        ws.cell(r, 4, pf["pv"]).number_format = DOLLAR
    _alt_rows(ws, r - len(val["pv_fcfs"]) + 1, r, 4)

    # ── Terminal value ────────────────────────────────────────────────────
    r += 2
    ws.cell(r, 1, "Terminal Value").font = SUB
    r += 1
    r = _lv(ws, r, "Terminal Year FCF", val["term_fcf"], DOLLAR)
    r = _lv(ws, r, "Terminal Growth Rate", scenario["tgr"], PCT)
    r = _lv(ws, r, "WACC", scenario["wacc"], PCT)
    r = _lv(ws, r, "Terminal Value (Gordon Growth)", val["tv"], DOLLAR)
    r = _lv(ws, r, "PV of Terminal Value", val["pv_tv"], DOLLAR)

    # ── EV to equity bridge ───────────────────────────────────────────────
    r += 1
    ws.cell(r, 1, "Enterprise → Equity Bridge").font = SUB
    r += 1
    r = _lv(ws, r, "Sum of PV of FCFs", val["sum_pv"], DOLLAR)
    r = _lv(ws, r, "+ PV of Terminal Value", val["pv_tv"], DOLLAR)
    r = _lv(ws, r, "= Enterprise Value", val["ev"], DOLLAR)
    r = _lv(ws, r, "- Net Debt", val["net_debt"], DOLLAR)
    r = _lv(ws, r, "= Equity Value", val["equity"], DOLLAR)
    r = _lv(ws, r, "÷ Shares Outstanding", val["shares"], NUM)
    r += 1
    ws.cell(r, 1, "Intrinsic Value Per Share").font = Font(bold=True, size=12,
                                                            color="1F3864", name="Calibri")
    c = ws.cell(r, 2, val["ivps"])
    c.font = Font(bold=True, size=12, color="1F3864", name="Calibri")
    c.number_format = DOLLAR2

    # ── Sensitivity table ─────────────────────────────────────────────────
    wacc_steps, tgr_steps, grid = sens_data
    r += 3
    ws.cell(r, 1, "Sensitivity Analysis: Implied Share Price").font = SUB
    r += 1
    ws.cell(r, 1, "WACC \\ TGR").font = HDR_FONT
    ws.cell(r, 1).fill = HDR_FILL
    for j, t in enumerate(tgr_steps):
        c = ws.cell(r, j + 2, t)
        c.number_format = PCT
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate(wacc_steps):
        r += 1
        c = ws.cell(r, 1, w)
        c.number_format = PCT
        c.fill = HDR_FILL
        c.font = HDR_FONT
        for j, price in enumerate(grid[i]):
            c = ws.cell(r, j + 2, price)
            c.number_format = DOLLAR2
            # Highlight the center cell
            if i == 2 and j == 2:
                c.fill = PatternFill("solid", fgColor="BDD7EE")
                c.font = BOLD

    _autowidth(ws)
    ws.column_dimensions["A"].width = 32
    ws.freeze_panes = "B4"


def sheet_ppe(wb, h, p):
    ws = wb.create_sheet("PPE & DA Schedule")
    hl, pl = _col_labels(h, p)
    ncols = 1 + len(hl) + len(pl)

    ws.cell(1, 1, "PP&E & Depreciation Schedule").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _hdr(ws, 3, [""] + hl + pl)

    # Historical PPE schedule
    h_ppe_begin = [0] + h["ppe_net"][:-1]
    h_capex = [abs(c) for c in h["capex"]]
    h_da = h["da"]
    h_ppe_end = h["ppe_net"]
    h_da_pct = [_pct(h_da[i], h_ppe_begin[i]) if h_ppe_begin[i] else 0
                for i in range(len(h["years"]))]

    lines = [
        ("Beginning PP&E", h_ppe_begin, p["ppe_begin"], DOLLAR),
        ("+ Capital Expenditure", h_capex, p["capex"], DOLLAR),
        ("- Depreciation & Amortisation", h_da, p["da"], DOLLAR),
        ("= Ending PP&E", h_ppe_end, p["ppe_end"], DOLLAR),
        ("", [], [], None),
        ("D&A as % of Beg. PP&E",
         h_da_pct,
         [_pct(p["da"][i], p["ppe_begin"][i]) for i in range(len(p["da"]))],
         PCT),
    ]

    r = 4
    for label, hv, pv, fmt in lines:
        if not label:
            r += 1
            continue
        ws.cell(r, 1, label).font = BOLD
        for j, v in enumerate(hv):
            ws.cell(r, j + 2, v).number_format = fmt or DOLLAR
        offset = len(hv) + 2
        for j, v in enumerate(pv):
            c = ws.cell(r, offset + j, v)
            c.number_format = fmt or DOLLAR
            c.fill = INPUT_FILL
        r += 1

    _alt_rows(ws, 4, r - 1, ncols)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_wc(wb, h, p):
    ws = wb.create_sheet("Working Capital Schedule")
    hl, pl = _col_labels(h, p)
    ncols = 1 + len(hl) + len(pl)

    ws.cell(1, 1, "Working Capital Schedule").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _hdr(ws, 3, [""] + hl + pl)

    # Compute historical NWC and delta
    h_nwc = [(h["receivables"][i] + h["inventory"][i] + h["prepaid"][i]
              - h["payables"][i] - h["accrued"][i]) for i in range(len(h["years"]))]
    h_delta = [0] + [h_nwc[i] - h_nwc[i - 1] for i in range(1, len(h_nwc))]

    nwc_items = [
        ("Revenue (reference)", "revenue", "revenue", DOLLAR, None),
        ("", None, None, None, None),
        ("Accounts Receivable", "receivables", "receivables", DOLLAR, "ar_pct"),
        ("  as % of Revenue", None, None, PCT, "ar_pct_row"),
        ("Inventory", "inventory", "inventory", DOLLAR, "inv_pct"),
        ("  as % of Revenue", None, None, PCT, "inv_pct_row"),
        ("Prepaid & Other", "prepaid", "prepaid", DOLLAR, "pre_pct"),
        ("  as % of Revenue", None, None, PCT, "pre_pct_row"),
        ("Accounts Payable", "payables", "payables", DOLLAR, "ap_pct"),
        ("  as % of Revenue", None, None, PCT, "ap_pct_row"),
        ("Accrued Expenses", "accrued", "accrued", DOLLAR, "acc_pct"),
        ("  as % of Revenue", None, None, PCT, "acc_pct_row"),
        ("", None, None, None, None),
        ("Net Working Capital", None, None, DOLLAR, "nwc_row"),
        ("Change in NWC", None, None, DOLLAR, "delta_row"),
    ]

    r = 4
    for label, h_key, p_key, fmt, tag in nwc_items:
        if not label:
            r += 1
            continue

        ws.cell(r, 1, label).font = BOLD if "%" not in label else Font(
            italic=True, size=9, color="666666", name="Calibri")

        if h_key and p_key:
            for j in range(len(h["years"])):
                ws.cell(r, j + 2, h[h_key][j]).number_format = fmt
            off = len(h["years"]) + 2
            for j in range(len(p["years"])):
                c = ws.cell(r, off + j, p[p_key][j])
                c.number_format = fmt
                c.fill = INPUT_FILL
        elif tag and "_pct_row" in tag:
            base = tag.replace("_pct_row", "")
            h_map = {"ar": "receivables", "inv": "inventory", "pre": "prepaid",
                     "ap": "payables", "acc": "accrued"}
            hk = h_map[base]
            for j in range(len(h["years"])):
                ws.cell(r, j + 2, _pct(h[hk][j], h["revenue"][j])).number_format = PCT
            off = len(h["years"]) + 2
            driver = p[f"{base}_pct"]
            for j in range(len(p["years"])):
                ws.cell(r, off + j, driver).number_format = PCT
        elif tag == "nwc_row":
            for j in range(len(h["years"])):
                ws.cell(r, j + 2, h_nwc[j]).number_format = DOLLAR
            off = len(h["years"]) + 2
            for j in range(len(p["years"])):
                c = ws.cell(r, off + j, p["nwc"][j])
                c.number_format = DOLLAR
                c.fill = INPUT_FILL
        elif tag == "delta_row":
            for j in range(len(h["years"])):
                ws.cell(r, j + 2, h_delta[j]).number_format = DOLLAR
            off = len(h["years"]) + 2
            for j in range(len(p["years"])):
                c = ws.cell(r, off + j, p["delta_nwc"][j])
                c.number_format = DOLLAR
                c.fill = INPUT_FILL
        r += 1

    _alt_rows(ws, 4, r - 1, ncols)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_wacc(wb, wc):
    ws = wb.create_sheet("WACC")
    ws.cell(1, 1, "Weighted Average Cost of Capital").font = TITLE
    ws.merge_cells("A1:B1")

    r = 3
    ws.cell(r, 1, "Cost of Equity (CAPM)").font = SUB
    r += 1
    r = _lv(ws, r, "Risk-Free Rate (10Y Treasury)", wc["rf"], PCT2)
    r = _lv(ws, r, "Beta (β)", wc["beta"], '0.00')
    r = _lv(ws, r, "Equity Risk Premium", wc["erp"], PCT2)
    r = _lv(ws, r, "Cost of Equity = Rf + β × ERP", wc["ke"], PCT2)
    r += 1
    ws.cell(r, 1, "Cost of Debt").font = SUB
    r += 1
    r = _lv(ws, r, "Interest Expense", wc["interest"], DOLLAR)
    r = _lv(ws, r, "Total Debt", wc["total_debt"], DOLLAR)
    r = _lv(ws, r, "Cost of Debt = Interest / Debt", wc["kd"], PCT2)
    r = _lv(ws, r, "Effective Tax Rate", wc["tax_rate"], PCT2)
    r = _lv(ws, r, "After-Tax Cost of Debt = Kd × (1 − t)", wc["kd_at"], PCT2)
    r += 1
    ws.cell(r, 1, "Capital Structure Weights").font = SUB
    r += 1
    r = _lv(ws, r, "Market Capitalisation", wc["mktcap"], DOLLAR)
    r = _lv(ws, r, "Total Debt", wc["total_debt"], DOLLAR)
    r = _lv(ws, r, "Total Capital", wc["total_cap"], DOLLAR)
    r = _lv(ws, r, "Equity Weight (We)", wc["we"], PCT2)
    r = _lv(ws, r, "Debt Weight (Wd)", wc["wd"], PCT2)
    r += 1
    ws.cell(r, 1, "WACC = We × Ke + Wd × Kd(1−t)").font = SUB
    r += 1
    c = ws.cell(r, 1, "WACC")
    c.font = Font(bold=True, size=12, color="1F3864", name="Calibri")
    c = ws.cell(r, 2, wc["wacc"])
    c.font = Font(bold=True, size=12, color="1F3864", name="Calibri")
    c.number_format = PCT2

    _autowidth(ws, mn=16, mx=40)
    ws.column_dimensions["A"].width = 38
    ws.freeze_panes = "A3"


def sheet_scenarios(wb, results, current_price):
    ws = wb.create_sheet("Scenarios")
    ws.cell(1, 1, "Scenario Analysis").font = TITLE
    ws.merge_cells("A1:D1")

    _hdr(ws, 3, ["", "Base", "Bull", "Bear"])

    rows = [
        ("Revenue Growth Rate", "rev_growth", PCT),
        ("WACC", "wacc", PCT),
        ("Terminal Growth Rate", "tgr", PCT),
        ("Intrinsic Value / Share", "ivps", DOLLAR2),
        ("Current Price", "price", DOLLAR2),
        ("Discount / Premium", "disc_pct", PCT),
        ("Verdict", "verdict", None),
    ]

    r = 4
    for label, key, fmt in rows:
        ws.cell(r, 1, label).font = BOLD
        for j, name in enumerate(["Base", "Bull", "Bear"]):
            res = results[name]
            val = res.get(key, "")
            c = ws.cell(r, j + 2, val)
            if fmt:
                c.number_format = fmt
            # Colour verdict cells
            if key == "verdict":
                c.alignment = Alignment(horizontal="center")
                if val == "UNDERVALUED":
                    c.fill, c.font = GREEN_FILL, GREEN_FNT
                elif val == "OVERVALUED":
                    c.fill, c.font = RED_FILL, RED_FNT
                else:
                    c.fill, c.font = YELLOW_FILL, YELLOW_FNT
        r += 1

    _alt_rows(ws, 4, r - 1, 4)
    _autowidth(ws, mn=18)
    ws.freeze_panes = "B4"


def sheet_comps(wb, ticker, stock_info, h):
    ws = wb.create_sheet("Comps")
    ws.cell(1, 1, "Comparable Company Analysis").font = TITLE
    ws.merge_cells("A1:H1")

    headers = ["Company", "Price", "Market Cap", "EV", "EBITDA",
               "EV/EBITDA", "P/E", "EV/Revenue"]
    _hdr(ws, 3, headers)

    mktcap = stock_info.get("marketCap", 0)
    price = stock_info.get("currentPrice") or stock_info.get("regularMarketPrice", 0)
    debt = (h["lt_debt"][-1] or 0) + (h["current_debt"][-1] or 0)
    cash = h["cash"][-1] or 0
    ev = mktcap + debt - cash
    ebitda = h["ebitda"][-1] or 1
    rev = h["revenue"][-1] or 1
    pe = stock_info.get("trailingPE", 0)

    r = 4
    ws.cell(r, 1, ticker).font = BOLD
    ws.cell(r, 2, price).number_format = DOLLAR2
    ws.cell(r, 3, mktcap).number_format = DOLLAR
    ws.cell(r, 4, ev).number_format = DOLLAR
    ws.cell(r, 5, ebitda).number_format = DOLLAR
    ws.cell(r, 6, ev / ebitda if ebitda else 0).number_format = MULT
    ws.cell(r, 7, pe).number_format = MULT
    ws.cell(r, 8, ev / rev if rev else 0).number_format = MULT

    # Blank peer rows
    for i in range(5):
        r += 1
        ws.cell(r, 1, "").font = NORM

    _alt_rows(ws, 4, r, 8)
    _autowidth(ws)
    ws.freeze_panes = "B4"


def sheet_summary(wb, ticker, results, h, wacc_comp, current_price, warns=None):
    ws = wb.create_sheet("Summary")
    ws.cell(1, 1, f"VALUS DCF — {ticker}").font = Font(bold=True, size=16,
                                                         color="1F3864", name="Calibri")
    ws.merge_cells("A1:C1")
    ws.cell(2, 1, f"Generated {date.today().isoformat()}").font = Font(
        size=9, color="888888", name="Calibri")

    r = 4
    ws.cell(r, 1, "Market Data").font = SUB
    r += 1
    r = _lv(ws, r, "Current Share Price", current_price, DOLLAR2)
    r = _lv(ws, r, "Date", date.today().isoformat())
    r += 1

    ws.cell(r, 1, "Intrinsic Value").font = SUB
    r += 1
    for name in ["Base", "Bull", "Bear"]:
        r = _lv(ws, r, f"  {name} Case", results[name]["ivps"], DOLLAR2)
    r += 1

    ws.cell(r, 1, "Key Assumptions").font = SUB
    r += 1
    r = _lv(ws, r, "WACC (Base)", results["Base"]["wacc"], PCT2)
    r = _lv(ws, r, "Terminal Growth (Base)", results["Base"]["tgr"], PCT2)
    r += 1

    ws.cell(r, 1, "Historical Metrics").font = SUB
    r += 1
    revs = [rv for rv in h["revenue"] if rv and rv > 0]
    rev_cagr = _cagr(revs[0], revs[-1], len(revs) - 1) if len(revs) >= 2 else 0
    ebitda_m = _pct(h["ebitda"][-1], h["revenue"][-1])
    # FCF margin
    fcf_vals = h.get("fcf")
    if not fcf_vals:
        # calculate from CFO - capex
        fcf_vals = [(h["cfo"][i] or 0) + (h["capex"][i] or 0) for i in range(len(h["years"]))]
    fcf_m = _pct(fcf_vals[-1] if fcf_vals else 0, h["revenue"][-1])
    r = _lv(ws, r, "Revenue CAGR", rev_cagr, PCT)
    r = _lv(ws, r, "EBITDA Margin (LTM)", ebitda_m, PCT)
    r = _lv(ws, r, "FCF Margin (LTM)", fcf_m, PCT)
    r += 1

    ws.cell(r, 1, "Verdict (Base Case)").font = SUB
    r += 1
    v = results["Base"]["verdict"]
    c = ws.cell(r, 1, v)
    if v == "UNDERVALUED":
        c.fill, c.font = GREEN_FILL, Font(bold=True, size=14, color="006100", name="Calibri")
    elif v == "OVERVALUED":
        c.fill, c.font = RED_FILL, Font(bold=True, size=14, color="9C0006", name="Calibri")
    else:
        c.fill, c.font = YELLOW_FILL, Font(bold=True, size=14, color="9C6500", name="Calibri")

    # ── Data warnings ────────────────────────────────────────────────────
    if warns:
        r += 2
        ws.cell(r, 1, "Data Warnings").font = SUB
        r += 1
        warn_font = Font(size=9, color="CC6600", name="Calibri")
        for w in warns:
            ws.cell(r, 1, f"  {w}").font = warn_font
            r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTABLE API
# ═══════════════════════════════════════════════════════════════════════════════

def default_scenarios(h, wacc):
    """Build default Base/Bull/Bear scenarios from historical data."""
    revs = [r for r in h["revenue"] if r and r > 0]
    hist_cagr = _cagr(revs[0], revs[-1], len(revs) - 1) if len(revs) >= 2 else 0.05

    return {
        "Base": dict(rev_growth=hist_cagr, wacc=wacc, tgr=0.025),
        "Bull": dict(rev_growth=hist_cagr * 1.5, wacc=max(wacc - 0.01, 0.05), tgr=0.03),
        "Bear": dict(rev_growth=max(hist_cagr * 0.5, 0.01), wacc=wacc + 0.01, tgr=0.02),
    }


def generate_dcf(ticker, scenarios=None):
    """
    Full DCF pipeline. Importable entry point for the web app.

    Args:
        ticker: Stock ticker string (e.g. "AAPL")
        scenarios: dict with Base/Bull/Bear keys, each having:
            rev_growth (decimal), wacc (decimal), tgr (decimal)
            If None, uses defaults from historical data.

    Returns:
        dict with keys: filename, excel_bytes, results, current_price,
                        wacc, wacc_comp, hist
    """
    from io import BytesIO

    h, stock, warns = fetch_all(ticker)
    info = stock.info
    current_price = info.get("currentPrice") or info.get("regularMarketPrice") or 0
    shares = info.get("sharesOutstanding") or 1

    # Set currency formats based on reporting currency
    currency_code = info.get("currency", "USD")
    _init_currency(currency_code)

    wacc_val, wacc_comp = calc_wacc(h, stock)

    if scenarios is None:
        scenarios = default_scenarios(h, wacc_val)

    all_results = {}
    base_proj = None
    base_val = None

    for name, sc in scenarios.items():
        sc["tax_rate"] = wacc_comp["tax_rate"]
        proj = build_projections(h, sc)
        proj["total_debt"] = (h["lt_debt"][-1] or 0) + (h["current_debt"][-1] or 0)
        proj["cash"] = h["cash"][-1] or 0

        val = run_valuation(proj, sc, shares)
        disc_pct = _pct(val["ivps"] - current_price, current_price)

        if disc_pct > 0.20:
            verdict = "UNDERVALUED"
        elif disc_pct < -0.20:
            verdict = "OVERVALUED"
        else:
            verdict = "FAIR"

        all_results[name] = {
            "rev_growth": sc["rev_growth"], "wacc": sc["wacc"], "tgr": sc["tgr"],
            "ivps": val["ivps"], "price": current_price,
            "disc_pct": disc_pct, "verdict": verdict,
            "proj": proj, "val": val,
        }
        if name == "Base":
            base_proj = proj
            base_val = val

    sens = sensitivity_table(base_val, base_proj, shares, scenarios["Base"]["wacc"])

    wb = Workbook()
    sheet_income(wb, h, base_proj)
    sheet_balance(wb, h)
    sheet_cashflow(wb, h)
    sheet_dcf(wb, h, base_proj, base_val, wacc_comp, sens, scenarios["Base"])
    sheet_ppe(wb, h, base_proj)
    sheet_wc(wb, h, base_proj)
    sheet_wacc(wb, wacc_comp)
    sheet_scenarios(wb, all_results, current_price)
    sheet_comps(wb, ticker, info, h)
    sheet_summary(wb, ticker, all_results, h, wacc_comp, current_price, warns)

    filename = f"DCF_{ticker}_{date.today().isoformat()}.xlsx"

    # Save to bytes for web download
    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    # Also save to disk
    wb.save(filename)

    return {
        "filename": filename,
        "excel_bytes": excel_bytes,
        "results": all_results,
        "current_price": current_price,
        "wacc": wacc_val,
        "wacc_comp": wacc_comp,
        "hist": h,
        "currency": currency_code,
        "currency_symbol": _CUR_SYMBOL,
        "warnings": warns,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 dcf.py <TICKER>")
        print("Example: python3 dcf.py AAPL")
        sys.exit(1)

    ticker = sys.argv[1].upper()
    print(f"\n{'=' * 56}")
    print(f"  VALUS Institutional DCF — {ticker}")
    print(f"{'=' * 56}\n")

    print("  Fetching historical financials...")
    h, stock, _ = fetch_all(ticker)
    wacc, _ = calc_wacc(h, stock)

    # Interactive prompts if running in a terminal, else defaults
    if sys.stdin.isatty():
        scenarios = get_scenarios(h, wacc)
    else:
        scenarios = default_scenarios(h, wacc)

    print("  Building model...")
    out = generate_dcf(ticker, scenarios)

    b = out["results"]["Base"]
    bu = out["results"]["Bull"]
    be = out["results"]["Bear"]
    price = out["current_price"]
    cur = out.get("currency_symbol", "$")
    sign = lambda x: "+" if x > 0 else ""

    print(f"\n{'=' * 56}")
    print(f"  {ticker} — DCF Summary ({out.get('currency', 'USD')})")
    print(f"{'=' * 56}")
    print(f"  Current Price:       {cur}{price:>10,.2f}")
    print(f"  ── Base Case ──")
    print(f"  Intrinsic Value:     {cur}{b['ivps']:>10,.2f}  ({sign(b['disc_pct'])}{b['disc_pct']:.1%})  {b['verdict']}")
    print(f"  ── Bull Case ──")
    print(f"  Intrinsic Value:     {cur}{bu['ivps']:>10,.2f}  ({sign(bu['disc_pct'])}{bu['disc_pct']:.1%})  {bu['verdict']}")
    print(f"  ── Bear Case ──")
    print(f"  Intrinsic Value:     {cur}{be['ivps']:>10,.2f}  ({sign(be['disc_pct'])}{be['disc_pct']:.1%})  {be['verdict']}")
    print(f"{'─' * 56}")
    print(f"  WACC (Base): {b['wacc']:.2%}  |  Terminal Growth: {b['tgr']:.1%}")
    print(f"  Saved: {out['filename']}")
    print(f"{'=' * 56}")

    if out.get("warnings"):
        print(f"\n  Warnings ({len(out['warnings'])}):")
        for w in out["warnings"]:
            print(f"    - {w}")
    print()


if __name__ == "__main__":
    main()
